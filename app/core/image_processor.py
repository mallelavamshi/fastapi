import os
import requests
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import tempfile
from datetime import datetime
import uuid
import gdown
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

def get_image_files(folder_path):
    """Get list of image files from folder"""
    valid_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')
    image_files = []
    
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(valid_extensions):
            image_files.append(filename)
    
    return image_files

def process_google_drive_link(link):
    """Process Google Drive link and download images"""
    temp_dir = tempfile.mkdtemp()
    downloaded_files = []
    
    try:
        if 'drive.google.com' in link:
            if 'folder' in link:
                # Handle folder link
                folder_id = link.split('/')[-1]
                if '?' in folder_id:
                    folder_id = folder_id.split('?')[0]
                gdown.download_folder(url=link, output=temp_dir, quiet=False)
                downloaded_files = [os.path.join(temp_dir, f) for f in os.listdir(temp_dir) 
                                 if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'))]
            else:
                # Handle single file link
                file_id = link.split('/')[-2]
                output = os.path.join(temp_dir, "downloaded_image.jpg")
                gdown.download(f"https://drive.google.com/uc?id={file_id}", output, quiet=False)
                if os.path.exists(output):
                    downloaded_files = [output]
    except Exception as e:
        print(f"Error downloading from Google Drive: {str(e)}")
        
    return temp_dir, downloaded_files

def upload_file(image_path, api_key):
    """Upload file to get upload_file_id"""
    upload_url = 'https://api.dify.ai/v1/files/upload'
    headers = {
        'Authorization': f'Bearer {api_key}'
    }
    
    with open(image_path, 'rb') as image_file:
        files = {
            'file': (os.path.basename(image_path), image_file, 'image/jpeg')
        }
        response = requests.post(upload_url, headers=headers, files=files)
        response.raise_for_status()
        return response.json().get('id')

def process_image(image_path, api_key):
    """Process single image through Dify AI API"""
    try:
        file_id = upload_file(image_path, api_key)
        
        url = 'https://api.dify.ai/v1/chat-messages'
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

        payload = {
            "inputs": {},
            "query": "Analyze this image and provide a detailed description and value assessment",
            "response_mode": "streaming",
            "conversation_id": "",
            "user": "abc-123",
            "files": [
                {
                    "type": "image",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id
                }
            ]
        }

        response = requests.post(url, headers=headers, json=payload, stream=True)
        response.raise_for_status()
        
        full_response = ""
        for line in response.iter_lines():
            if line:
                try:
                    line_text = line.decode('utf-8')
                    if line_text.startswith('data: '):
                        line_text = line_text[6:]
                    
                    data = json.loads(line_text)
                    
                    if data.get('event') == 'agent_message' and 'answer' in data:
                        full_response += data['answer']
                        
                except (json.JSONDecodeError, Exception):
                    continue
        
        return full_response

    except requests.exceptions.RequestException as e:
        return f"Error processing image: {str(e)}"

def process_image_for_excel(image_path, cell_height):
    """Process and resize image for Excel with dynamic height"""
    try:
        with Image.open(image_path) as img:
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            # Fixed width, dynamic height
            max_width = 200
            
            # Calculate new dimensions maintaining aspect ratio
            aspect_ratio = img.width / img.height
            new_width = min(max_width, img.width)
            new_height = int(min(cell_height, new_width / aspect_ratio))
            
            # Resize image
            resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            img_byte_arr = io.BytesIO()
            resized_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            
            return img_byte_arr
            
    except Exception as e:
        print(f"Error processing image {image_path}: {str(e)}")
        return None

def create_excel_with_images(results, output_dir):
    """Create Excel file with images and analysis"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = str(uuid.uuid4())[:8]
    excel_filename = f'analysis_results_{timestamp}_{unique_id}.xlsx'
    excel_path = os.path.join(output_dir, excel_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"
    
    # Set column headers
    ws['A1'] = "Image"
    ws['B1'] = "Image Name"
    ws['C1'] = "Analysis"
    
    # Style headers
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
    
    # Set column widths
    ws.column_dimensions['A'].width = 30  # ~200 pixels
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30  # ~200 pixels
    
    row = 2
    for result in results:
        # Process analysis text
        analysis_text = result['API_Response']
        ws.cell(row=row, column=2, value=result['Image_Name'])
        analysis_cell = ws.cell(row=row, column=3, value=analysis_text)
        
        # Set text wrapping and alignment
        analysis_cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calculate required height for text
        approx_chars_per_line = 30  # Based on column width
        num_lines = len(analysis_text) / approx_chars_per_line
        text_height = max(75, min(400, num_lines * 15))  # 15 pixels per line, min 75px, max 400px
        
        # Set row height based on content
        row_height = max(text_height, 200)  # At least 200 pixels for image
        ws.row_dimensions[row].height = row_height * 0.75  # Convert pixels to points
        
        try:
            img_data = process_image_for_excel(result['Image_Path'], row_height)
            if img_data:
                img = XLImage(img_data)
                ws.add_image(img, f'A{row}')
            else:
                ws.cell(row=row, column=1, value="Error processing image")
        except Exception as e:
            print(f"Error adding image to Excel: {str(e)}")
            ws.cell(row=row, column=1, value="Error loading image")
        
        row += 1
    
    wb.save(excel_path)
    return excel_path

def convert_to_pdf(excel_path, output_dir):
    """Convert results to PDF using ReportLab"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = str(uuid.uuid4())[:8]
    pdf_filename = f'analysis_results_{timestamp}_{unique_id}.pdf'
    pdf_path = os.path.join(output_dir, pdf_filename)
    
    # Read Excel file
    df = pd.read_excel(excel_path)
    
    # Create PDF
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=30*mm,
        leftMargin=30*mm,
        topMargin=30*mm,
        bottomMargin=30*mm
    )
    
    # Prepare story (content)
    story = []
    styles = getSampleStyleSheet()
    
    # Add title
    title = Paragraph("Image Analysis Results", styles['Heading1'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Process each row
    for index, row in df.iterrows():
        # Add image name
        image_name = Paragraph(f"Image: {row['Image Name']}", styles['Heading2'])
        story.append(image_name)
        story.append(Spacer(1, 10))
        
        # Add analysis text
        analysis = Paragraph(row['Analysis'], styles['Normal'])
        story.append(analysis)
        story.append(Spacer(1, 20))
        
        # Add a line separator
        story.append(Paragraph("_" * 50, styles['Normal']))
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    
    return pdf_path